#!/usr/bin/python
import datetime
import csv
import re
import sys
import unicodedata
from openpyxl import Workbook

def correctDate(d):
	try:
		dt = re.search(r'(\d+)\/(\d+)\/(\d+)',d)
		if dt:
			d = dt.group(1)
			m = dt.group(2)
			y = dt.group(3)
			return datetime.date(int(y),int(m),int(d))
	except:
		print "failed date"

def correctDialed(dial):
	try:
		#if re.search(r'(\D+)',dial):
		#	print dial
		#	dial = re.sub('(i)','',dial)
		d = re.search(r'(\d+)',dial)
		if d:
			return str(d.group(1))
	except:
		print "failed dialed"

def correctTime(t):
	try:
		tx = re.search(r'(\d+)\:(\d+)\:(\d+)',t)
		if tx:
			h = tx.group(1)
			m = tx.group(2)
			s = tx.group(3)
			return datetime.time(int(h),int(m),int(s))
	except:
		print "failed time"
			
def correctDur(d):
	try:
		if int(d) < 86399:
			durh = int(int(d)/3600)
			durm = int((int(d)- (int(int(d) /3600)*3600))/60) 
			durs = int(d) % 60
			return datetime.time(durh,durm,durs)
	except:
		print "failed duration"
		return datetime.time(00,00,00)

def correctType(t):
	try:
		if re.search(r'GPRS',t):
                	return "Data"
		else:
			return "Voice"
	except:
		print "failed type"
		return "Voice"

def cap(s,l):
	return s if len(s)<=l else s[0:l]

def correctCost(c):
	try:
		co = re.sub(',','.',c)
		return float(co)
	except:
		print "failed cost"
		return float("0")

def correctRoam(r):
	try:
		if re.search(r'%',r):
			return "True"
		else:
			return "False" 
	except:
		print "failed roam"
			
def correctDir(d):
		return "Outgoing"	
		
def writeHeaders(ws):
	ws.cell(row=0,column=0).value = "Date"
	ws.cell(row=0,column=1).value = "Time"
	ws.cell(row=0,column=2).value = "Asset Number"
	ws.cell(row=0,column=3).value = "Dialed Number"
	ws.cell(row=0,column=4).value = "Duration"
	ws.cell(row=0,column=5).value = "Cost"
	ws.cell(row=0,column=6).value = "Type"
	ws.cell(row=0,column=7).value = "Destination"
	ws.cell(row=0,column=8).value = "Roaming"
	ws.cell(row=0,column=9).value = "Direction"
	ws.cell(row=0,column=10).value = "Data Usage"
	ws.cell(row=0,column=11).value = "URL"
	ws.cell(row=0,column=12).value = "Currency"
	
def main():
	csvfile = raw_input('Enter csv file you want to convert: ')
	c = 1 
	wb = Workbook()
	xl = re.sub('\.','-',csvfile)	
	dest = xl + '-converted.xlsx'
	ws = wb.active
	ws.title = "turkcell"
	writeHeaders(ws)
	with open(csvfile,'rU') as fin:
		try:
			print "opening turkcell.csv"
			reader = csv.reader(fin)
			for index,row in enumerate(reader):
				if re.search(r'CUST',row[0]) is not None:
					continue
				if re.search(r'^\d+',row[0]) is not None:
					dat = correctDate(row[4])
					tim = correctTime(row[4])
					ws.cell(row=c,column=0).value = dat # date
					ws.cell(row=c,column=1).value = tim # time
					asset = cap(row[1],20)
					ws.cell(row=c,column=2).value = asset # asset number
					dialed = correctDialed(row[2])
					if dialed: dialed = cap(dialed,20)
					ws.cell(row=c,column=3).value = dialed # dialed number
					ws.cell(row=c,column=3).style.number_format.format_code = '0'
					dur = correctDur(row[6])
					ws.cell(row=c,column=4).value = dur # duration
					cost = correctCost(row[8])
					ws.cell(row=c,column=5).value = cost # cost
					typ = correctType(row[5])
					ws.cell(row=c,column=6).value = typ # type
					
					dstn = cap(row[5],25)
					dstn = dstn.decode('ascii','ignore')
					dstn = dstn.encode('ascii','ignore')
					ws.cell(row=c,column=7).value = dstn # destination
					roam = correctRoam(row[4])
					ws.cell(row=c,column=8).value = roam # roaming
					direction = correctDir(row[4])
					ws.cell(row=c,column=9).value = direction # direction
					ws.cell(row=c,column=10).value = row[7] # data usage
					#url = cap(row[5],250)
					url = ""
					ws.cell(row=c,column=11).value = url # url
					ws.cell(row=c,column=12).value = 'TRY' # currency
					c += 1
				else:
					print "line error", c
					pass
		except UnicodeDecodeError:
			print "unicode error"
		except:
			print "failed main loop"
			#print c
			
			print sys.exc_info()[0]

	print "saving " + dest
	wb.save(filename = dest)
	
if __name__ == "__main__":
	sys.exit(main())
	#main()

