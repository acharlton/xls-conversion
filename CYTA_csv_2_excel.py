#!/usr/bin/python
import datetime
import csv
import re
import sys
from openpyxl import Workbook

def correctDate(d):
	date = re.search(r'(\d+)\-(\d+)\-(\d+)',d)
	if date:
		d = date.group(1)
		m = date.group(2)
		y = date.group(3)
		return datetime.date(int(y),int(m),int(d))

def check_exclusion(dial,desc,cost):
	# return True if we want to exclude this call
	# the majority of calls will be local ie to number starting with 22,99 etc try to match them first
	if not dial:
		# if no characters in dialed number field
		return True
	if len(dial) <= 8:
		# small numbers are local services
		return True
	if cost == 0:
		# exclude zero cost calls
		return True

	# go though exclusion list
	try:
		lst = ['^357','^00357','^90','^0090']
		for i,prefix in enumerate(lst):
			print "checking for prefix: ", prefix, "with dialed number" , dial
			d = re.search(prefix,dial)
			if d:
				#print "desc ",desc
				print "found ", prefix, "in ", dial
				de = desc.find('%')
				if de > -1:
					# is roaming call so present for verification
					print "found roaming call, processing: ", desc, " at position: ", de
					return False
				else:
					print "Not roaming" 
					return True	
			else:
				print "No match for: ", prefix
		return False	
	except:
		print "exception found during exclusion check: ",dial
		print sys.exc_info()[0]
		return True

def correctDialed(dial):
	d = re.search(r'(\d+)',dial)
	if d:
		return str(d.group(1))

def correctTime(t):
	tx = re.search(r'(\d+)\:(\d+)\:(\d+)',t)
	if tx:
		h = tx.group(1)
		m = tx.group(2)
		s = tx.group(3)
		return datetime.time(int(h),int(m),int(s))
			
def correctDur(d):
	if int(d) < 86399:
		durh = int(int(d)/3600)
		durm = int((int(d)- (int(int(d) /3600)*3600))/60) 
		durs = int(d) % 60
		return datetime.time(durh,durm,durs)

def correctType(t):
	if re.search(r'V',t):
		return "Voice"
	if re.search(r'S',t):
		return "SMS"
	if re.search(r'D',t):
		return "Data"
	if re.search(r'G',t):
		return "Data"
	else:
		return "Voice"

def cap(s,l):
	return s if len(s)<=l else s[0:l]

def correctCost(c):
	co = re.sub(',','.',c)
	return float(co)

def correctRoam(r):
	if re.search(r'%',r):
		return "True"
	else:
		return "False" 
			
def correctDir(d):
	if re.search(r'O',d):
		return "Outgoing"
	elif re.search(r'I',d):
		return "Incoming"
	elif re.search(r'F',d):
		return "Outgoing"
	else:
		return ""	
		
def writeHeaders(ws):
	ws.cell(row=0,column=0).value = "Date"
	ws.cell(row=0,column=1).value = "Time"
	ws.cell(row=0,column=2).value = "Asset Number"
	ws.cell(row=0,column=3).value = "Dialed Number"
	ws.cell(row=0,column=4).value = "Duration"
	ws.cell(row=0,column=5).value = "Cost"
	ws.cell(row=0,column=6).value = "Type"
	ws.cell(row=0,column=7).value = "Destination"
	ws.cell(row=0,column=8).value = "Roaming"
	ws.cell(row=0,column=9).value = "Direction"
	ws.cell(row=0,column=10).value = "Data Usage"
	ws.cell(row=0,column=11).value = "URL"
	ws.cell(row=0,column=12).value = "Currency"
	
def main():
	c = 1
	l = 0 
	wb = Workbook()
	mo = raw_input('Please specify the 2 digit month of the calls.csv file: ')
	dest = r'cyta_converted_' + mo + '.xlsx'
	ws = wb.active
	ws.title = "cyta"
	writeHeaders(ws)
	with open('calls.csv','rU') as fin:
		try:
			print "opening calls.csv"
			reader = csv.reader(fin,delimiter=';')
			for index,row in enumerate(reader):
				l += 1
				print "starting ",l
				check = re.match('CALL',row[0])
				if check:
					# called_num,description,amount
					cost = correctCost(row[12])
					exclude = check_exclusion(row[15],row[11],cost)
					if exclude:
						print "excluding: ", "line ", l, row[15]
					else:
						print "processing ", l
						dat = correctDate(row[17])
						tim = correctTime(row[17])
						ws.cell(row=c,column=0).value = dat # date
						ws.cell(row=c,column=1).value = tim # time
						asset = cap(row[7],20)
						ws.cell(row=c,column=2).value = asset # asset number
						dialed = correctDialed(row[15])
						if dialed: dialed = cap(dialed,20)
						ws.cell(row=c,column=3).value = dialed # dialed number
						ws.cell(row=c,column=3).style.number_format.format_code = '0'
						dur = correctDur(row[19])
						ws.cell(row=c,column=4).value = dur # duration
						cost = correctCost(row[12])
						ws.cell(row=c,column=5).value = cost # cost
						typ = correctType(row[22])
						ws.cell(row=c,column=6).value = typ # type
						dstn = cap(row[11],25)
						ws.cell(row=c,column=7).value = dstn # destination
						roam = correctRoam(row[11])
						ws.cell(row=c,column=8).value = roam # roaming
						direction = correctDir(row[21])
						ws.cell(row=c,column=9).value = direction # direction
						ws.cell(row=c,column=10).value = row[20] # data usage
						url = cap(row[11],250)
						ws.cell(row=c,column=11).value = url # url
						ws.cell(row=c,column=12).value = 'EUR' # currency
						c += 1
		except:
			print "failed opening csv file"
			print sys.exc_info()[0]

	print "saving " + dest
	wb.save(filename = dest)
	sheet = wb.worksheets[0]
	row_count = sheet.get_highest_row() - 1
	print "Total records converted: " ,row_count
	
if __name__ == "__main__":
	#sys.exit(main())
	main()

